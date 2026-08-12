"""
db_manager.py - Модуль управления базой данных аккаунтов для Spotify Automation Manager

Этот модуль работает с файлом accounts.xlsx и предоставляет функции для:
- Проверки существования файла БД и создания его с заголовками
- Чтения всех аккаунтов из файла
- Обновления статуса конкретного аккаунта
"""

import os
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise ImportError("Требуется установить openpyxl: pip install openpyxl")


@dataclass
class Account:
    """Класс представления аккаунта Spotify"""
    profile_id: str           # ID профиля (например, spotify_001)
    manager_name: str         # Название/Имя менеджера (например, Account_DE)
    proxy: str                # Прокси в формате IP:PORT:USER:PASS
    login: str                # Логин Spotify
    password: str             # Пароль Spotify
    playlist_url: str         # Ссылка на целевой плейлист
    start_track_url: str      # Ссылка на стартовый трек для прослушивания
    status: str = "Спит"      # Статус (Спит / В работе / Ошибка / Готово)
    
    def to_dict(self) -> Dict[str, str]:
        """Конвертирует объект Account в словарь"""
        return asdict(self)
    
    @classmethod
    def from_dict(cls, data: Dict[str, str]) -> 'Account':
        """Создает объект Account из словаря"""
        return cls(
            profile_id=data.get('profile_id', ''),
            manager_name=data.get('manager_name', ''),
            proxy=data.get('proxy', ''),
            login=data.get('login', ''),
            password=data.get('password', ''),
            playlist_url=data.get('playlist_url', ''),
            start_track_url=data.get('start_track_url', ''),
            status=data.get('status', 'Спит')
        )


# Заголовки колонок в файле Excel
COLUMN_HEADERS = [
    "ID профиля",
    "Название/Имя менеджера",
    "Прокси (IP:PORT:USER:PASS)",
    "Логин Spotify",
    "Пароль Spotify",
    "Ссылка на целевой плейлист",
    "Ссылка на стартовый трек",
    "Статус"
]

# Маппинг имен полей класса Account на номера колонок (0-indexed)
FIELD_TO_COLUMN = {
    'profile_id': 0,
    'manager_name': 1,
    'proxy': 2,
    'login': 3,
    'password': 4,
    'playlist_url': 5,
    'start_track_url': 6,
    'status': 7
}


class DatabaseManager:
    """Менеджер базы данных аккаунтов Spotify"""
    
    def __init__(self, db_path: str = "accounts.xlsx", profiles_dir: str = "profiles"):
        """
        Инициализирует менеджер базы данных.
        
        Args:
            db_path: Путь к файлу базы данных Excel
            profiles_dir: Путь к директории для хранения профилей (куки)
        """
        self.db_path = Path(db_path)
        self.profiles_dir = Path(profiles_dir)
        
        # Создаем директорию для профилей если она не существует
        self.profiles_dir.mkdir(parents=True, exist_ok=True)
        
        # Проверяем и создаем файл БД если необходимо
        self._ensure_db_exists()
    
    def _ensure_db_exists(self) -> None:
        """
        Проверяет существование файла базы данных.
        Если файл не существует, создает его с правильными заголовками.
        """
        if not self.db_path.exists():
            self._create_empty_database()
    
    def _create_empty_database(self) -> None:
        """Создает новый файл базы данных с заголовками колонок"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"
        
        # Добавляем заголовки
        for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Сохраняем файл
        wb.save(str(self.db_path))
        wb.close()
    
    def _load_workbook(self) -> tuple:
        """
        Загружает рабочую книгу и активный лист.
        
        Returns:
            Tuple[Workbook, Worksheet]: Кортеж из workbook и worksheet
        """
        wb = load_workbook(str(self.db_path))
        ws = wb.active
        return wb, ws
    
    def read_all_accounts(self) -> List[Account]:
        """
        Читает все аккаунты из файла базы данных.
        
        Returns:
            List[Account]: Список объектов Account
        """
        if not self.db_path.exists():
            return []
        
        wb, ws = self._load_workbook()
        accounts = []
        
        # Начинаем со второй строки (первая - заголовки)
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            
            # Читаем данные из строки
            row_data['profile_id'] = str(ws.cell(row=row_idx, column=1).value or '')
            row_data['manager_name'] = str(ws.cell(row=row_idx, column=2).value or '')
            row_data['proxy'] = str(ws.cell(row=row_idx, column=3).value or '')
            row_data['login'] = str(ws.cell(row=row_idx, column=4).value or '')
            row_data['password'] = str(ws.cell(row=row_idx, column=5).value or '')
            row_data['playlist_url'] = str(ws.cell(row=row_idx, column=6).value or '')
            row_data['start_track_url'] = str(ws.cell(row=row_idx, column=7).value or '')
            row_data['status'] = str(ws.cell(row=row_idx, column=8).value or 'Спит')
            
            # Пропускаем пустые строки
            if not row_data['profile_id'] and not row_data['manager_name']:
                continue
            
            account = Account.from_dict(row_data)
            accounts.append(account)
        
        wb.close()
        return accounts
    
    def update_account_status(self, profile_id: str, new_status: str) -> bool:
        """
        Обновляет статус конкретного аккаунта в файле.
        
        Args:
            profile_id: ID профиля аккаунта для обновления
            new_status: Новый статус для установки
            
        Returns:
            bool: True если обновление прошло успешно, False если аккаунт не найден
        """
        if not self.db_path.exists():
            return False
        
        wb, ws = self._load_workbook()
        updated = False
        
        # Ищем аккаунт по ID профиля
        for row_idx in range(2, ws.max_row + 1):
            current_id = str(ws.cell(row=row_idx, column=1).value or '')
            
            if current_id == profile_id:
                # Обновляем статус в колонке 8
                ws.cell(row=row_idx, column=8, value=new_status)
                updated = True
                break
        
        if updated:
            wb.save(str(self.db_path))
        
        wb.close()
        return updated
    
    def get_account_by_id(self, profile_id: str) -> Optional[Account]:
        """
        Получает аккаунт по его ID.
        
        Args:
            profile_id: ID профиля для поиска
            
        Returns:
            Optional[Account]: Объект Account если найден, иначе None
        """
        accounts = self.read_all_accounts()
        for account in accounts:
            if account.profile_id == profile_id:
                return account
        return None
    
    def get_profile_path(self, profile_id: str) -> Path:
        """
        Получает путь к директории профиля для хранения куки.
        
        Args:
            profile_id: ID профиля
            
        Returns:
            Path: Путь к директории профиля
        """
        profile_path = self.profiles_dir / profile_id
        profile_path.mkdir(parents=True, exist_ok=True)
        return profile_path
    
    def add_account(self, account: Account) -> bool:
        """
        Добавляет новый аккаунт в базу данных.
        
        Args:
            account: Объект Account для добавления
            
        Returns:
            bool: True если добавление прошло успешно
        """
        wb, ws = self._load_workbook()
        
        # Находим следующую пустую строку
        next_row = ws.max_row + 1
        
        # Записываем данные аккаунта
        ws.cell(row=next_row, column=1, value=account.profile_id)
        ws.cell(row=next_row, column=2, value=account.manager_name)
        ws.cell(row=next_row, column=3, value=account.proxy)
        ws.cell(row=next_row, column=4, value=account.login)
        ws.cell(row=next_row, column=5, value=account.password)
        ws.cell(row=next_row, column=6, value=account.playlist_url)
        ws.cell(row=next_row, column=7, value=account.start_track_url)
        ws.cell(row=next_row, column=8, value=account.status)
        
        wb.save(str(self.db_path))
        wb.close()
        return True
    
    def get_accounts_count(self) -> int:
        """
        Получает общее количество аккаунтов в базе данных.
        
        Returns:
            int: Количество аккаунтов
        """
        return len(self.read_all_accounts())
    
    def get_accounts_by_status(self, status: str) -> List[Account]:
        """
        Получает список аккаунтов с указанным статусом.
        
        Args:
            status: Статус для фильтрации
            
        Returns:
            List[Account]: Список аккаунтов с указанным статусом
        """
        accounts = self.read_all_accounts()
        return [acc for acc in accounts if acc.status == status]


# Пример использования
if __name__ == "__main__":
    # Инициализация менеджера БД
    db = DatabaseManager()
    
    print(f"Файл БД: {db.db_path.absolute()}")
    print(f"Директория профилей: {db.profiles_dir.absolute()}")
    print(f"Количество аккаунтов: {db.get_accounts_count()}")
    
    # Чтение всех аккаунтов
    accounts = db.read_all_accounts()
    print(f"\nСписок аккаунтов ({len(accounts)}):")
    for acc in accounts:
        print(f"  - {acc.profile_id}: {acc.manager_name} [{acc.status}]")
    
    # Пример обновления статуса
    if accounts:
        test_id = accounts[0].profile_id
        print(f"\nОбновление статуса для {test_id}...")
        success = db.update_account_status(test_id, "В работе")
        print(f"Результат: {'Успешно' if success else 'Ошибка'}")
        
        # Возвращаем статус обратно для демонстрации
        db.update_account_status(test_id, "Спит")
